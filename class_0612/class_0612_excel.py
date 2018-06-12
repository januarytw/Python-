__author__ = 'Administrator'
# 操作excel
# 安装pip install openpyxl
# excel 工作簿workbook---表单sheet----单元格cell


from openpyxl import load_workbook  #只能操作xlsx的文件

# 读取值
# 1、打开工作簿
work_book=load_workbook("case.xlsx")#


# 2、确定表单
sheet=work_book["test_data"]
# sheet=work_book.get_sheet_by_name('test_data')

# 3、确定单元格位置
print(sheet.cell(2,6).value)

# 4、写入值
sheet.cell(2,7).value='asdasdf'
sheet.cell(3,7).value='我是一个额结算单爱的'
work_book.save("case.xlsx")#保存

# sheet.max_row 统计行数
# sheet.max_column 统计列数
