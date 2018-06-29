'''
完成注册/登录/充值接口的请求：（每个接口写3条用例，至少保证有2条是正常的数据）
接口地址示范（登录）：119.23.241.154:8080/futureloan/mvc/api/member/login

1：请自己把测试数据写到Excel里面去
2：利用我们之前写好的Excel读取和存储测试数据的类，从1中获取必要的测试数据
3：完成接口的测试
4：同步把测试结果存到Excel中。
5：写代码之前，先自己思考怎么写。
6：附加作业：请利用思维导图工具xmind画出你写代码的思路以及调用关系（要写!!!)
'''

# 操作excel类
from openpyxl import load_workbook
class DoExcel():
    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name

    #读取Excel数据
    def readData(self):#此处可以改进一下，利用初始化函数
        try:
            work_book=load_workbook(self.file_path)
            sheet=work_book[self.sheet_name]
            i=1
            list_1=[]
            list_2=[]
            while i<sheet.max_row:#!!!!此处可以用for in range()
                j=1
                while j<=sheet.max_column:
                    list_1.append(sheet.cell(i+1,j).value)
                    j+=1
                i+=1
                list_2.append(list_1)
                list_1=[]
            return list_2
        except Exception as e:
            print("读取出错了：",e)

    #写入excel
    def writeData(self,r,c,msg):
        try:
            work_book=load_workbook(self.file_path)
            sheet=work_book[self.sheet_name]
            sheet.cell(r,c).value=msg
            work_book.save(self.file_path)#保存
        except Exception as e:
            print("写入出错了：",e)

#请求接口类
import requests

doExcel=DoExcel('case.xlsx','test_data')
excel_result=doExcel.readData()

#如果需要登录操作时，调用此session
url='http://119.23.241.154:8080/futureloan/mvc/api/member/login'
data={'mobilephone':'13141222222','pwd':'123456'}
s=requests.session()
s.post(url,data)

for item in excel_result:
    if item[1]=="注册":
        try:
            #注册
            login_url=item[4]
            login_data=eval(item[5])
            result=requests.get(login_url,login_data).json()
            doExcel.writeData(item[0]+1,8,str(result))
        except Exception as e:
            print ("请求失败，出现的错误是%s"%e)
            raise e
    elif item[1]=="登录":
        try:
            login_url=item[4]
            login_data=eval(item[5])
            print(login_data)
            result_login=requests.post(login_url,login_data).json()
            # print(result_login)
            doExcel.writeData(item[0]+1,8,str(result_login))
        except Exception as e:
            print ("请求失败，出现的错误是%s"%e)
            raise e
    elif item[1]=="充值":
        try:
            #充值
            login_url=item[4]
            login_data=eval(item[5])
            result_recharge=s.post(login_url,login_data).json()
            # print(result_recharge)
            doExcel.writeData(item[0]+1,8,str(result_recharge))
        except Exception as e:
            print ("请求失败，出现的错误是%s"%e)
            raise e

