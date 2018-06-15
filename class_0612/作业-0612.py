
# 作业：0612
# 1：url地址：http://v.juhe.cn/laohuangli/d
# 请求参数：{'date':'2018-09-11','key':'a8f2732319cf0ad3cce8ec6ef7aa4f33'}
# 2：请根据上课内容对老黄历编写至少5条用例，上述给出的数据都是正确的请求数据
# 3：把用例的数据写到Excel里面，然后请把每一行的数据存到一个子列表里面，所有的行数据都放到一个大的列表里面。
# 4：读取数据的的功能，请写成一个类。
<<<<<<< .mine
import
=======

>>>>>>> .theirs
from openpyxl import load_workbook
class ReadData():
    #读取Excel数据，并把每一行的数据追加到列表中
    def ReadExcelToList(self,file):
        try:
            work_book=load_workbook(file)
            sheet=work_book["test_data"]##工作表的名字可以作为参数
            row=sheet.max_row
            cloumn=sheet.max_column

            i=1
            list_1=[]
            list_2=[]
            while i<row:#!!!!此处可以用for in range()
                j=1
                while j<=cloumn:
                    list_1.append(sheet.cell(i+1,j).value)
                    j+=1
                i+=1
                list_2.append(list_1)
                list_1=[]
            return list_2
        except Exception as e:
            print("出错了：",e)


print(ReadData().ReadExcelToList("case.xlsx"))
