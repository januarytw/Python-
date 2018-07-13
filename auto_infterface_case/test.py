from openpyxl import load_workbook
from openpyxl.styles import Font,colors,Fill,PatternFill,

wb=load_workbook(r'D:\Python34\TestCode\auto_infterface_case\test_data\test_case.xlsx')
sheet=wb['init']
