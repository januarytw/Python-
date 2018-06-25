# 1：利用6-12号写的excel数据读取类以及6月9号写的HTTP单元测试类，完成老黄历的单元测试，且输出测试报告。
# （注意是要从Excel里面获取测试 数据哦~并且把测试结果写回到Excel中去。请注意听老师上课的讲解题意。）


'''一、HTTP请求类 ,分别完成GET和POST请求'''
import requests
class HttpRequest():
    def __init__(self,url,data=None):#空用None
        self.url=url
        self.data=data

    def get_post_request(self,method):
        try:
            if method.upper()=="GET":
                response=requests.get(self.url,self.data)
                return response.json()
            elif method.upper()=="POST":
                response=requests.post(self.url,self.data)
                return response.json()
        except Exception as e:
            print ("请求失败，出现的错误是%s"%e)
            raise e



'''二、操作excel'''
from openpyxl import load_workbook
class DoExcel():
    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name

    #读取Excel数据
    def readData(self):#此处可以改进一下，利用初始化函数
        try:
            work_book=load_workbook(self.file_path)
            sheet=work_book[self.sheet_name]
            i=1
            list_1=[]
            list_2=[]
            while i<sheet.max_row:#!!!!此处可以用for in range()
                j=1
                while j<=sheet.max_column:
                    list_1.append(sheet.cell(i+1,j).value)
                    j+=1
                i+=1
                list_2.append(list_1)
                list_1=[]
            return list_2
        except Exception as e:
            print("读取出错了：",e)

    #写入excel
    def writeData(self,r,c,msg):
        try:
            work_book=load_workbook(self.file_path)
            sheet=work_book[self.sheet_name]
            sheet.cell(r,c).value=msg
            work_book.save(self.file_path)#保存
        except Exception as e:
            print("写入出错了：",e)



'''三、http请求类的单元测试'''
import unittest
from ddt import ddt,data,unpack

excelResult=DoExcel("case.xlsx","test_data").readData()
@ddt
class TestHttpRequest(unittest.TestCase):
    def setUp(self):
        self.doExcel=DoExcel("case.xlsx","test_data")
        print("测试开始！")
        # print(excelResult)

    @data(*excelResult)
    def test_http_request(self,data_1):
        print("执行的是第",data_1[0],"条用例")
        response=HttpRequest(url=data_1[3],data=eval(data_1[4])).get_post_request(data_1[2])#！！！data参数是字典型的，所以要用eval()转换
        try:
            self.assertEqual(data_1[5],response["reason"])#正常就通过，结果为否时，抛出异常
            result='pass'
        except AssertionError as e:
            print("出错了",e)
            result='fail'
            raise e
        finally:
            #写入数据
            self.doExcel.writeData(data_1[0]+1,7,str(response))#！！！写入时，要将字典类型转换成字符串
            self.doExcel.writeData(data_1[0]+1,8,result)


    def tearDown(self):
        print("测试结束！\n")


if __name__=="__main__":
    import time
    import HTMLTestRunnerNew
    suite=unittest.TestSuite()
    # suite.addTests(TestHttpRequest('test_http_request'))
    loader=unittest.TestLoader()
    suite.addTests(loader.loadTestsFromTestCase(TestHttpRequest))

    now = time.strftime('%Y-%m-%d_%H_%M_%S')#获取当前时间!!!
    file_path="test"+now+".html"
    with open(file_path,"wb") as file:
        runner=HTMLTestRunnerNew.HTMLTestRunner(stream=file, verbosity=2,title='测试报告',description='作文大全接口测试',tester='张三')
        runner.run(suite)
