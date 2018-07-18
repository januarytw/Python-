'''
操作excel，读取和写入操作
'''

from openpyxl import load_workbook
from openpyxl.styles import Font,colors
from conf import project_path
from common.read_config import ReadConfig

class DoExcel():
    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name

    #读取为未注册的手机号
    def no_reg_tel(self):
        wb=load_workbook(self.file_path)
        sheet=wb['init']
        no_reg_tel=sheet.cell(1,2).value
        return no_reg_tel

    #更新未注册的手机号，手机号码+1
    def update_tel(self,new_tel):
        wb=load_workbook(self.file_path)
        sheet=wb['init']
        sheet.cell(1,2).value=new_tel
        wb.save(self.file_path)

    #读取excel中的用例,返回的数据类型为list
    def read_data_to_list(self,mode,case_list):#mode：读取用例模式，0为指定用例，1为全部用例。case_list：用例列表
        #载入文件
        wb=load_workbook(self.file_path)
        #指定工作表
        sheet=wb[self.sheet_name]
        #实现每一行数据存在一个列表里面，然后所有行的数据存在一个大列表里面
        test_data=[]#存储所有行的数据
        no_reg_tel=self.no_reg_tel()
        if mode=='1':#读取全部的用例
            for i in range(2,sheet.max_row+1):#从excel的第2行开始读；range取左不取右，所以要+1
                sub_data=[]#存储每一行的数据
                for j in range(1,9):
                    if j==6:
                        param=eval(sheet.cell(i,6).value)
                        # print('param=eval(sheet.cell(i+1,6).value)的值为：%s,类型为%s'%(param,type(param)))
                        if param['mobilephone']=='first_tel':
                            param['mobilephone']=no_reg_tel
                        sub_data.append(param)
                    elif j==8:
                        check_sql=eval(sheet.cell(i,8).value)
                        if check_sql['sql_data']=='first_tel':
                            check_sql['sql_data']=no_reg_tel
                        sub_data.append(check_sql)
                    else:
                        sub_data.append(sheet.cell(i,j).value)
                test_data.append(sub_data)
        elif mode=='0':
            for i in case_list:#！！！此处需要case_list是list类型的，从配置文件里读取出来要进行转换
                sub_data=[]#存储每一行的数据
                for j in range(1,8):
                    if j==6:
                        param=eval(sheet.cell(i+1,6).value)#i+1的意思 ：是从第二行开始的
                        if param["mobilephone"]=='first_tel':
                            param['mobilephone']=no_reg_tel
                        sub_data.append(param)#如果有first_tel这个就替换后加到列表中，如果没有就直接加入
                    elif j==8:
                        check_sql=eval(sheet.cell(i,8).value)
                        if check_sql['sql_data']=='first_tel':
                            check_sql['sql_data']=no_reg_tel
                        sub_data.append(check_sql)
                    else:
                        sub_data.append(sheet.cell(i+1,j).value)
                test_data.append(sub_data)
        self.update_tel(str(int(no_reg_tel)+1))

        return test_data

    #读取excel中的用例,返回的数据类型为dict
    def read_data(self,mode,case_list):#mode：读取用例模式，0为指定用例，1为全部用例。case_list：用例列表
        #载入文件
        wb=load_workbook(self.file_path)
        #指定工作表
        sheet=wb[self.sheet_name]
        #实现每一行数据存在一个列表里面，然后所有行的数据存在一个大列表里面
        test_data=[]#存储所有行的数据
        no_reg_tel=self.no_reg_tel()

        #方法一
        # if mode=='1':
        #     for i in range(2,sheet.max_row+1):
        #         sub_data={}
        #         sub_data['case_id']=sheet.cell(i,1).value
        #         sub_data['method']=sheet.cell(i,4).value
        #         sub_data['url']=sheet.cell(i,5).value
        #         sub_data['expect_result']=sheet.cell(i,7).value
        #
        #         #对请求参数param进行变量替换
        #         if sheet.cell(i,6).value.find('${no_reg_tel}')!=-1:
        #             new_param=sheet.cell(i,6).value.replace('${no_reg_tel}',no_reg_tel)
        #         else:
        #             new_param=sheet.cell(i,6).value
        #         sub_data['params']=new_param
        #
        #         #对check_sql进行变量替换
        #         if sheet.cell(i,8).value!=None and sheet.cell(i,8).value.find('${no_reg_tel}')!=-1:
        #             new_param=sheet.cell(i,8).value.replace('${no_reg_tel}',no_reg_tel)
        #         else:
        #             new_param=sheet.cell(i,8).value
        #         sub_data['params']=new_param
        #         test_data.append(sub_data)
        #
        # elif mode=='0':
        #     for i in case_list:
        #         sub_data={}#存到一个字典里面sub_data={}
        #         sub_data['case_id']=sheet.cell(i+1,1).value
        #         sub_data['method']=sheet.cell(i+1,4).value
        #         sub_data['url']=sheet.cell(i+1,5).value
        #         sub_data['expect_result']=sheet.cell(i+1,7).value
        #
        #         #对请求参数param进行替换
        #         if sheet.cell(i+1,6).value.find('${no_reg_tel}')!=-1:
        #            new_param=sheet.cell(i+1,6).value.replace('${no_reg_tel}',str(no_reg_tel))
        #         else:
        #            new_param=sheet.cell(i+1,6).value
        #         sub_data['params']=new_param
        #
        #         #对check_sql去进行更新值
        #         if sheet.cell(i+1,8).value.find('${fno_reg_tel}')!=-1:
        #            new_param=sheet.cell(i+1,8).value.replace('${no_reg_tel}',str(no_reg_tel))
        #         else:
        #            new_param=sheet.cell(i+1,8).value
        #         sub_data['check_sql']=new_param
        #         test_data.append(sub_data)
        #
        # self.update_tel(str(int(no_reg_tel)+1))
        #
        # return test_data

        #方法二
        for i in range(2,sheet.max_row+1):
            sub_data={}#存到一个字典里面sub_data={}
            sub_data['case_id']=sheet.cell(i+1,1).value
            sub_data['method']=sheet.cell(i+1,4).value
            sub_data['url']=sheet.cell(i+1,5).value
            sub_data['expect_result']=sheet.cell(i+1,7).value
            #对参数param进行替换
            if sheet.cell(i,6).value.find('${no_reg_tel}')!=-1:
                new_param=sheet.cell(i,6).value.replace('${no_reg_tel}',str(no_reg_tel))
            else:
                new_param=sheet.cell(i,6).value
            sub_data['params']=new_param

            #对check_sql去进行更新值
            if sheet.cell(i,8).value!=None and sheet.cell(i,8).value.find('${no_reg_tel}')!=-1:
                new_param=sheet.cell(i,8).value.replace('${no_reg_tel}',str(no_reg_tel))
            else:
                new_param=sheet.cell(i,8).value
            sub_data['check_sql']=new_param

            test_data.append(sub_data)#所有的数据都存在test_data里面

        if mode=='1':
            final_data=test_data
        else:
            final_data=[]
            for item in test_data:
                if item['case_id'] in case_list:
                    final_data.append(item)

        self.update_tel(str(int(no_reg_tel)+1))

        return final_data


    #写入excel
    def write_data(self,row,mode,actual,result):
        #载入文件
        wb=load_workbook(self.file_path)
        #指定工作表
        sheet=wb[self.sheet_name]
        #写入测试数据时，固定写入第8和9列
        if mode==1:
            sheet.cell(row,9).value=actual
            sheet.cell(row,10).value=result
        elif mode==2:
            sheet.cell(row,11).value=actual
            sheet.cell(row,12).value=result
        #保存
        wb.save(self.file_path)


if __name__ == '__main__':
    mode=ReadConfig(project_path.case_conf_path).getConfig('CASE','mode')
    case_list=eval(ReadConfig(project_path.case_conf_path).getConfig('CASE','case_list'))
    print (type(case_list))
    testdata=DoExcel(project_path.test_data_path,"test_data").read_data(mode,case_list)
    print(testdata)

