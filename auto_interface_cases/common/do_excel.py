'''
操作excel，读取和写入操作
'''

from openpyxl import load_workbook

class DoExcel():
    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name

    def read_data(self):
        #载入文件
        wb=load_workbook(self.file_path)
        #指定工作表
        sheet=wb[self.sheet_name]

        #实现每一行数据存在一个列表里面，然后所有行的数据存在一个大列表里面
        test_data=[]#存储所有行的数据
        for i in range(2,sheet.max_row+1):#从excel的第2行开始读；range取左不取右，所以要+1
            sub_data=[]#存储每一行的数据
            for j in range(1,8):
                sub_data.append(sheet.cell(i,j).value)
            test_data.append(sub_data)
        return test_data

    def write_data(self,row,actual,result):
        #载入文件
        wb=load_workbook(self.file_path)
        #指定工作表
        sheet=wb[self.sheet_name]

        #写入测试数据时，固定写入第8和9列
        sheet.cell(row,8).value=actual
        sheet.cell(row,9).value=result

        #保存
        wb.save(self.file_path)


if __name__ == '__main__':
    testdata=DoExcel('test_case.xlsx',"test_data").read_data()
    print(testdata)

